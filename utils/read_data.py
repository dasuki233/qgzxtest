# -*- coding: UTF-8 -*-
'''
@Project ：Python_Workspace 
@File    ：parse_excel.py
@IDE     ：PyCharm 
@Author  ：zjj
@Date    ：2023/8/21 20:35 
'''
import os.path

import openpyxl
import yaml

from qgzxtest.utils.logger import logger


class ReadData:

    def get_objectt_path(self):
        return os.path.dirname(__file__).split('utils')[0]

    def read_excel(self, path):
        path = self.get_objectt_path() + '/data/' + path
        logger.info("加载 {} 文件......".format(path))
        # path = self.get_objectt_path() + '/data/login_data.xlsx'
        # 加载excel
        wb = openpyxl.load_workbook(path)
        # 获得sheet对象
        sheet = wb['login']
        # 获得excel行列数
        rows, columns = sheet.max_row, sheet.max_column
        values_list = []
        for row in range(2, rows + 1):
            rows_list = []
            for column in range(1, columns + 1):
                rows_list.append(sheet.cell(row, column).value)
            values_list.append(rows_list)
            logger.info("读到数据 ==>>  {} ".format(rows_list))
        return values_list

    def load_yaml(self, path):
        path = self.get_objectt_path() + '/data/' + path
        logger.info("加载 {} 文件......".format(path))
        with open(path, encoding='utf-8') as f:
            data = yaml.safe_load(f)
        logger.info("读到数据 ==>>  {} ".format(data))
        return data


if __name__ == '__main__':
    rd = ReadData()
    data = rd.load_yaml('D:\Python_Workspace\qgzxtest\data\login_data.yml')
    print(data['test_normal_login'])
